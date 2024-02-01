import time
import json
import uuid


from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

import fcntl

def acquire_lock(file_path):
    file_handle = open(file_path, 'r')
    try:
        fcntl.flock(file_handle, fcntl.LOCK_EX | fcntl.LOCK_NB)
        print(f"Lock acquired on {file_path}")
    except IOError:
        print(f"Unable to acquire lock on {file_path}")
        return None
    return file_handle

def release_lock(file_handle):
    fcntl.flock(file_handle, fcntl.LOCK_UN)
    file_handle.close()
    print("Lock released")
    
    

def flatten_dict(d, parent_key='', sep='.'):
    """
    Flatten a nested dictionary by combining keys with separators.
    """
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v)) 
    return dict(items)


class HttpTraceMiddleware:
    def __init__(self, get_response):
        self.get_response = get_response

    def __call__(self, request):
        # Check if the trace ID is present in the request headers
        trace_id = request.headers.get('X-Trace-ID')

        # If not present, generate a new trace ID
        if not trace_id:
            trace_id = str(uuid.uuid4())

        # Generate span ID for the current request
        span_id = str(uuid.uuid4())

        # Capture the start time of the request
        start_time = time.time()

        response = self.get_response(request)

        # Capture the end time of the request
        end_time = time.time()

        # Calculate the request duration
        duration = int((end_time - start_time) * 1000)  # Convert to milliseconds

        # Add or update the trace ID in the response headers
        response['X-Trace-ID'] = trace_id

        # Get IP address
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        ip = x_forwarded_for.split(',')[0] if x_forwarded_for else request.META.get('REMOTE_ADDR')

        # Get user session hash
        session = request.session
        session_hash = session.get('_auth_user_hash', "NON LOGGED IN USER")

        # Get username
        username = request.user.username if request.user.is_authenticated else "NON LOGGED IN USER"

        # Create a dictionary with the trace log information
        trace_log = {
            "traceId": trace_id,
            "spanId": span_id,
            "timestamp": time.strftime('%Y-%m-%dT%H:%M:%S.000Z', time.gmtime()),
            "principal": username,  # Customize based on your authentication system
            "session": session_hash, 
            "username": username,
            "ip": ip,
            "request": {
                "method": request.method,
                "uri": request.build_absolute_uri(),
                "headers": dict(request.headers),
                "remoteAddress": ip,  # Customize based on your needs
            },
            "response": {
                "status": response.status_code,
                "headers": dict(response.headers),
            },
            "timeTaken": duration,
        }
        
        trace_log = flatten_dict(trace_log)
        
        file_path = "server_logs_new.xlsx"
        
        """
            Create a new Excel file if it doesn't exist, and add the trace log as the first row.
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            headers = list(trace_log.keys())
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            wb.save(file_path)
        
        #locking the file till the data is being written in a single process
        lock_handle = acquire_lock(file_path)

        if lock_handle:
            try:
                

                for value in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    headers = list(value)

                missing_columns = list(set(trace_log.keys()) - set(headers))
                if len(missing_columns):
                    last_col_index = ws.max_column

                    for column in missing_columns:
                        ws.cell(row=1, column=last_col_index + 1, value=column)
                        
                headers = headers + missing_columns
                for key, value in trace_log.items():
                # Find the column index based on the key
                    column_index = headers.index(key) + 1

                    # Find the last row with a value in the target column
                    last_row = ws.max_row
                    while last_row > 1 and ws.cell(row=last_row, column=column_index).value is None:
                        last_row -= 1

                    # Insert the value in the next available row
                    ws.cell(row=last_row + 1, column=column_index, value=value)

                wb.save(file_path)
            finally:
                release_lock(lock_handle)
            
            
            
            

        # try:
        #     with open('logger_new.json', 'r') as file:
        #         existing_logs = json.load(file)
        # except (FileNotFoundError, json.JSONDecodeError):
        #     existing_logs = []

        # # Combine existing logs with new logs
        # combined_logs = existing_logs + trace_log

        # # Write the combined logs back to the file
        # with open('logger_new.json', 'w') as file:
        #     file.write(json.dumps(combined_logs, indent=2))
        #     file.write('\n')

        return response
